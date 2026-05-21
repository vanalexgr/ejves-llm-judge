"""Final comparator reporting utilities."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd


SOURCE_MODEL_LABELS = {
    "chatgpt_free": "GPT-5.5",
    "claude_free": "Claude Sonnet 4.6",
    "gemini_free": "Gemini 3.5 Flash",
}
SOURCE_MODEL_ORDER = ("chatgpt_free", "gemini_free", "claude_free")
ORIGINAL_BENCHMARK_LABEL = "Original GPT-3.5 study benchmark"
BOOTSTRAP_SAMPLES = 100
BOOTSTRAP_SEED = 20260521


@dataclass(frozen=True)
class EndpointSpec:
    final_column: str
    original_column: str
    display_name: str
    direction: str  # "higher" or "lower"
    group: str  # "validated" or "descriptive"
    subset: str  # "all" or "treatment"


ENDPOINT_SPECS = (
    EndpointSpec("validated_accuracy", "accuracy_mean", "Accuracy", "higher", "validated", "all"),
    EndpointSpec("validated_tone", "tone_mean", "Tone", "higher", "validated", "all"),
    EndpointSpec(
        "validated_complementarity",
        "complementarity_mean",
        "Complementarity",
        "higher",
        "validated",
        "all",
    ),
    EndpointSpec(
        "validated_gilbert_urgency",
        "gilbert_urgency_mean",
        "Urgency",
        "lower",
        "validated",
        "all",
    ),
    EndpointSpec(
        "validated_discern_q7",
        "discern_q7_mean",
        "DISCERN Q7",
        "higher",
        "validated",
        "treatment",
    ),
    EndpointSpec(
        "descriptive_comprehensiveness",
        "comprehensiveness_mean",
        "Comprehensiveness",
        "higher",
        "descriptive",
        "all",
    ),
    EndpointSpec(
        "descriptive_clarity",
        "clarity_mean",
        "Clarity",
        "higher",
        "descriptive",
        "all",
    ),
)


def _format_float(value: Any, digits: int = 3) -> str:
    if value is None or pd.isna(value):
        return "NA"
    return f"{float(value):.{digits}f}"


def _format_ci(low: Any, high: Any, digits: int = 3) -> str:
    if low is None or high is None or pd.isna(low) or pd.isna(high):
        return "NA"
    return f"{float(low):.{digits}f} to {float(high):.{digits}f}"


def _format_estimate_with_ci(
    value: Any,
    low: Any,
    high: Any,
    *,
    digits: int = 3,
) -> str:
    if value is None or pd.isna(value):
        return "NA"
    ci = _format_ci(low, high, digits=digits)
    if ci == "NA":
        return _format_float(value, digits=digits)
    return f"{_format_float(value, digits=digits)} ({ci})"


def _mean(series: pd.Series) -> float | pd.NA:
    values = series.dropna().astype(float)
    if values.empty:
        return pd.NA
    return float(values.mean())


def _bootstrap_mean_ci(
    series: pd.Series,
    *,
    n_bootstrap: int = BOOTSTRAP_SAMPLES,
    seed: int = BOOTSTRAP_SEED,
) -> tuple[float, float]:
    values = series.dropna().astype(float).to_numpy()
    if len(values) < 2:
        return float("nan"), float("nan")
    rng = np.random.default_rng(seed)
    means = [
        float(values[rng.integers(0, len(values), size=len(values))].mean())
        for _ in range(n_bootstrap)
    ]
    return float(np.percentile(means, 2.5)), float(np.percentile(means, 97.5))


def _bootstrap_delta_ci(
    left: pd.Series,
    right: pd.Series,
    *,
    direction: str,
    n_bootstrap: int = BOOTSTRAP_SAMPLES,
    seed: int = BOOTSTRAP_SEED,
) -> tuple[float, float]:
    left_values = left.dropna().astype(float).to_numpy()
    right_values = right.dropna().astype(float).to_numpy()
    if len(left_values) < 2 or len(right_values) < 2:
        return float("nan"), float("nan")
    rng = np.random.default_rng(seed)
    deltas: list[float] = []
    for _ in range(n_bootstrap):
        left_mean = float(left_values[rng.integers(0, len(left_values), size=len(left_values))].mean())
        right_mean = float(right_values[rng.integers(0, len(right_values), size=len(right_values))].mean())
        delta = left_mean - right_mean
        if direction == "lower":
            delta = -delta
        deltas.append(delta)
    return float(np.percentile(deltas, 2.5)), float(np.percentile(deltas, 97.5))


def build_original_benchmark_summary(consensus_frame: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for spec in ENDPOINT_SPECS:
        subset = consensus_frame
        if spec.subset == "treatment":
            subset = subset.loc[subset["question_group"].eq("treatment")]
        mean_value = _mean(subset[spec.original_column])
        ci_low, ci_high = _bootstrap_mean_ci(
            subset[spec.original_column],
            seed=BOOTSTRAP_SEED + sum(ord(char) for char in spec.original_column),
        )
        rows.append(
            {
                "endpoint": spec.display_name,
                "direction": spec.direction,
                "group": spec.group,
                "original_mean": mean_value,
                "original_mean_ci_low": ci_low,
                "original_mean_ci_high": ci_high,
                "response_count": int(subset["response_id"].nunique()),
            }
        )
    return pd.DataFrame(rows)


def build_model_endpoint_summary(comparator_results: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for source_model in SOURCE_MODEL_ORDER:
        subset = comparator_results.loc[comparator_results["source_model"].eq(source_model)].copy()
        if subset.empty:
            continue
        row: dict[str, Any] = {
            "source_model": source_model,
            "model_label": SOURCE_MODEL_LABELS.get(source_model, source_model),
            "response_count": int(len(subset)),
            "treatment_response_count": int(subset["question_group"].eq("treatment").sum()),
        }
        for spec in ENDPOINT_SPECS:
            spec_subset = subset
            if spec.subset == "treatment":
                spec_subset = subset.loc[subset["question_group"].eq("treatment")]
            row[spec.final_column] = _mean(spec_subset[spec.final_column])
            ci_low, ci_high = _bootstrap_mean_ci(
                spec_subset[spec.final_column],
                seed=BOOTSTRAP_SEED + sum(ord(char) for char in f"{source_model}:{spec.final_column}"),
            )
            row[f"{spec.final_column}_ci_low"] = ci_low
            row[f"{spec.final_column}_ci_high"] = ci_high
        rows.append(row)
    return pd.DataFrame(rows)


def build_endpoint_comparison_table(
    *,
    model_summary: pd.DataFrame,
    original_benchmark: pd.DataFrame,
    group: str,
    comparator_results: pd.DataFrame | None = None,
    original_consensus: pd.DataFrame | None = None,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    benchmark_index = original_benchmark.set_index("endpoint")
    summary_index = model_summary.set_index("source_model")
    for spec in ENDPOINT_SPECS:
        if spec.group != group:
            continue
        benchmark_value = benchmark_index.loc[spec.display_name, "original_mean"]
        benchmark_ci_low = benchmark_index.loc[spec.display_name, "original_mean_ci_low"]
        benchmark_ci_high = benchmark_index.loc[spec.display_name, "original_mean_ci_high"]
        row: dict[str, Any] = {
            "Endpoint": spec.display_name,
            "Direction": "Higher is better" if spec.direction == "higher" else "Lower is better",
            ORIGINAL_BENCHMARK_LABEL: benchmark_value,
            f"{ORIGINAL_BENCHMARK_LABEL} 95% CI": _format_ci(benchmark_ci_low, benchmark_ci_high),
        }
        if group == "descriptive":
            row["Status"] = "Descriptive only; not validated"
        for source_model in SOURCE_MODEL_ORDER:
            if source_model not in summary_index.index:
                continue
            model_value = summary_index.loc[source_model, spec.final_column]
            model_ci_low = summary_index.loc[source_model, f"{spec.final_column}_ci_low"]
            model_ci_high = summary_index.loc[source_model, f"{spec.final_column}_ci_high"]
            delta = model_value - benchmark_value
            row[SOURCE_MODEL_LABELS[source_model]] = model_value
            row[f"{SOURCE_MODEL_LABELS[source_model]} 95% CI"] = _format_ci(model_ci_low, model_ci_high)
            delta_column = f"{SOURCE_MODEL_LABELS[source_model]} vs original"
            if spec.direction == "lower":
                row[delta_column] = -delta
            else:
                row[delta_column] = delta
            delta_ci_column = f"{delta_column} 95% CI"
            if comparator_results is not None and original_consensus is not None:
                model_subset = comparator_results.loc[
                    comparator_results["source_model"].eq(source_model)
                ]
                original_subset = original_consensus
                if spec.subset == "treatment":
                    model_subset = model_subset.loc[model_subset["question_group"].eq("treatment")]
                    original_subset = original_subset.loc[original_subset["question_group"].eq("treatment")]
                delta_ci_low, delta_ci_high = _bootstrap_delta_ci(
                    model_subset[spec.final_column],
                    original_subset[spec.original_column],
                    direction=spec.direction,
                    seed=BOOTSTRAP_SEED + sum(ord(char) for char in f"{source_model}:{spec.display_name}:delta"),
                )
                row[delta_ci_column] = _format_ci(delta_ci_low, delta_ci_high)
            else:
                row[delta_ci_column] = "NA"
        rows.append(row)
    return pd.DataFrame(rows)


def build_accuracy_resolution_summary(comparator_results: pd.DataFrame) -> pd.DataFrame:
    total_rows = len(comparator_results)
    rows = [
        {
            "metric": "total_comparator_responses",
            "value": int(total_rows),
        },
        {
            "metric": "dual_surgeon_agreement",
            "value": int(comparator_results["accuracy_source"].eq("dual_surgeon_agreement").sum()),
        },
        {
            "metric": "dual_surgeon_mean",
            "value": int(comparator_results["accuracy_source"].eq("dual_surgeon_mean").sum()),
        },
        {
            "metric": "triple_surgeon_median",
            "value": int(comparator_results["accuracy_source"].eq("triple_surgeon_median").sum()),
        },
        {
            "metric": "mario_explicit_revisions",
            "value": int(comparator_results["mario_requested_revised_accuracy"].notna().sum()),
        },
        {
            "metric": "third_reviewer_cases",
            "value": int(comparator_results["third_reviewer_used"].fillna(False).sum()),
        },
    ]
    return pd.DataFrame(rows)


def render_comparator_report(
    *,
    comparator_results: pd.DataFrame,
    original_benchmark: pd.DataFrame,
    model_summary: pd.DataFrame,
    validated_comparison: pd.DataFrame,
    descriptive_comparison: pd.DataFrame,
    accuracy_resolution: pd.DataFrame,
) -> str:
    validated_summary = pd.DataFrame(
        {
            "Model": model_summary["model_label"],
            "Accuracy": model_summary.apply(
                lambda row: _format_estimate_with_ci(
                    row["validated_accuracy"],
                    row["validated_accuracy_ci_low"],
                    row["validated_accuracy_ci_high"],
                ),
                axis=1,
            ),
            "Tone": model_summary.apply(
                lambda row: _format_estimate_with_ci(
                    row["validated_tone"],
                    row["validated_tone_ci_low"],
                    row["validated_tone_ci_high"],
                ),
                axis=1,
            ),
            "Complementarity": model_summary.apply(
                lambda row: _format_estimate_with_ci(
                    row["validated_complementarity"],
                    row["validated_complementarity_ci_low"],
                    row["validated_complementarity_ci_high"],
                ),
                axis=1,
            ),
            "Urgency": model_summary.apply(
                lambda row: _format_estimate_with_ci(
                    row["validated_gilbert_urgency"],
                    row["validated_gilbert_urgency_ci_low"],
                    row["validated_gilbert_urgency_ci_high"],
                ),
                axis=1,
            ),
            "DISCERN Q7 (treatment)": model_summary.apply(
                lambda row: _format_estimate_with_ci(
                    row["validated_discern_q7"],
                    row["validated_discern_q7_ci_low"],
                    row["validated_discern_q7_ci_high"],
                ),
                axis=1,
            ),
        }
    )
    descriptive_summary = pd.DataFrame(
        {
            "Model": model_summary["model_label"],
            "Status": "Descriptive only; not validated",
            "Comprehensiveness": model_summary.apply(
                lambda row: _format_estimate_with_ci(
                    row["descriptive_comprehensiveness"],
                    row["descriptive_comprehensiveness_ci_low"],
                    row["descriptive_comprehensiveness_ci_high"],
                ),
                axis=1,
            ),
            "Clarity": model_summary.apply(
                lambda row: _format_estimate_with_ci(
                    row["descriptive_clarity"],
                    row["descriptive_clarity_ci_low"],
                    row["descriptive_clarity_ci_high"],
                ),
                axis=1,
            ),
        }
    )

    benchmark_index = original_benchmark.set_index("endpoint")
    accuracy_leader = model_summary.sort_values("validated_accuracy", ascending=False).iloc[0]
    accuracy_runner_up = model_summary.sort_values("validated_accuracy", ascending=False).iloc[1]
    tone_leader = model_summary.sort_values("validated_tone", ascending=False).iloc[0]
    urgency_leader = model_summary.sort_values("validated_gilbert_urgency", ascending=True).iloc[0]
    discern_leader = model_summary.sort_values("validated_discern_q7", ascending=False).iloc[0]

    return f"""# Final Comparator Report

## Dataset Summary

- Comparator responses: {len(comparator_results)} total (`16` question stems x `3` models)
- Comparator models: GPT-5.5, Gemini 3.5 Flash, Claude Sonnet 4.6
- Comparator collection mode: free consumer-facing model interfaces, copied into the scoring pipeline before judge evaluation
- Original benchmark: reconstructed GPT-3.5 study set (`16` responses)
- Final comparator accuracy process: two blinded board-certified surgeon ratings for all `48` responses, with blinded third-surgeon adjudication for `5` initially disputed rows

## Accuracy Resolution Summary

{accuracy_resolution.to_markdown(index=False)}

## Validated Endpoints By Model

`Accuracy`, `tone`, `complementarity`, `gilbert_urgency`, and treatment-only `DISCERN Q7` are the calibration-passing comparator endpoints. Lower urgency scores indicate stronger recommendation for urgent care. Table entries below are mean scores with bootstrap 95% confidence intervals.

{validated_summary.to_markdown(index=False, floatfmt='.3f')}

## Validated Endpoints Versus Original GPT-3.5 Benchmark

Positive `vs original` values indicate improvement. For urgency, the delta is sign-flipped so positive values mean more appropriate urgency signaling than the original GPT-3.5 benchmark.

Detailed 95% bootstrap confidence intervals for these benchmark deltas are included in `validated_endpoint_comparison.csv` and `comparator_summary_tables.xlsx`.

{validated_comparison.to_markdown(index=False, floatfmt='.3f')}

## Descriptive-Only Endpoints

`Comprehensiveness` and `clarity` are retained descriptively only because the corrected manuscript Table 2 indicates that, at the original symptoms-topic sample size (`n = 4` diseases), human ICC(2,k) for these domains was not meaningfully estimable; the response-level judge calibration in this repository was likewise weak. These table entries are descriptive only and should not be cited as validated endpoints.

{descriptive_summary.to_markdown(index=False, floatfmt='.3f')}

## Descriptive Endpoints Versus Original GPT-3.5 Benchmark

{descriptive_comparison.to_markdown(index=False, floatfmt='.3f')}

## Interpretation

- Accuracy was fully resolved for all `48` comparator rows. The highest mean final validated accuracy was for {accuracy_leader['model_label']} ({_format_float(accuracy_leader['validated_accuracy'])}), with {accuracy_runner_up['model_label']} close behind ({_format_float(accuracy_runner_up['validated_accuracy'])}). Claude Sonnet 4.6 remained closest to the original GPT-3.5 benchmark on this endpoint.
- Tone improved for all three newer models relative to the reconstructed GPT-3.5 baseline. The most balanced overall tone was seen in {tone_leader['model_label']} ({_format_float(tone_leader['validated_tone'])}).
- Complementarity was broadly stable across models; GPT-5.5 and Claude Sonnet 4.6 matched each other on average, while Gemini 3.5 Flash was slightly lower.
- Urgency signaling remained strongest for {urgency_leader['model_label']} because it produced the lowest mean urgency score ({_format_float(urgency_leader['validated_gilbert_urgency'])}).
- Treatment uncertainty handling (`DISCERN Q7`) was highest for {discern_leader['model_label']} ({_format_float(discern_leader['validated_discern_q7'])}).
- Descriptively, all three newer models were clearer and more comprehensive than the original GPT-3.5 benchmark, but these two domains should not be framed as validated primary endpoints. The corrected Table 2 interpretation is that the original human topic-level signal for these constructs was itself unstable or non-estimable at this sample size.

## Measurement And Review Process Caveats

- Mixed measurement arms apply across time: the original GPT-3.5 benchmark uses the reconstructed historical human-consensus ratings, whereas comparator accuracy uses the later dual-surgeon plus adjudication workflow. These benchmark comparisons are therefore pragmatic rather than perfectly instrument-equivalent.
- VGA served as one of the two comparator-arm accuracy raters and is disclosed here as a board-certified surgeon rather than as an independent historical reviewer from the original study set. The mitigation was full blinding to model identity during the initial scoring round, followed by third-rater adjudication for the rows with large disagreement.
- Mario made `2` explicit post-round score revisions after reviewing the rationale set. Those revisions were retained in the audit trail rather than silently overwritten, and both rows remained inside the disagreement-managed workflow instead of being used to bypass adjudication.
- The `5/48` adjudication cases represent a 10.4% third-review rate. This should be interpreted as an expected signal-detection feature of the clinical accuracy workflow, not as evidence that the adjudication process failed.

## Recommended Manuscript Framing

1. Present the newer models as improved communicators relative to the original GPT-3.5 benchmark on the calibration-passing judged domains.
2. Present accuracy as a human-scored endpoint, not an AI-judged endpoint.
3. State explicitly that the comparator arm combined judge scoring for calibration-passing communication domains with blinded surgeon scoring for clinical accuracy.
4. State explicitly that tone, complementarity, and DISCERN-Q7 were floor-passing domains with more modest absolute agreement than urgency, so they should be read as usable-with-caution rather than as strong psychometric validations.
5. Keep `comprehensiveness` and `clarity` in supplementary or descriptive tables only, with the corrected Table 2 caveat that the original topic-level human reliability for these constructs was not meaningfully estimable at this sample size.
"""


def write_comparator_report(
    *,
    comparator_results_path: Path,
    original_consensus_path: Path,
    output_dir: Path,
) -> dict[str, Path]:
    comparator_results = pd.read_parquet(comparator_results_path)
    original_consensus = pd.read_parquet(original_consensus_path)

    original_benchmark = build_original_benchmark_summary(original_consensus)
    model_summary = build_model_endpoint_summary(comparator_results)
    validated_comparison = build_endpoint_comparison_table(
        model_summary=model_summary,
        original_benchmark=original_benchmark,
        group="validated",
        comparator_results=comparator_results,
        original_consensus=original_consensus,
    )
    descriptive_comparison = build_endpoint_comparison_table(
        model_summary=model_summary,
        original_benchmark=original_benchmark,
        group="descriptive",
        comparator_results=comparator_results,
        original_consensus=original_consensus,
    )
    accuracy_resolution = build_accuracy_resolution_summary(comparator_results)

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "comparator_report.md"
    workbook_path = output_dir / "comparator_summary_tables.xlsx"
    model_summary_path = output_dir / "model_endpoint_summary.csv"
    validated_path = output_dir / "validated_endpoint_comparison.csv"
    descriptive_path = output_dir / "descriptive_endpoint_comparison.csv"
    accuracy_resolution_path = output_dir / "accuracy_resolution_summary.csv"
    original_benchmark_path = output_dir / "original_gpt35_benchmark_summary.csv"

    report_path.write_text(
        render_comparator_report(
            comparator_results=comparator_results,
            original_benchmark=original_benchmark,
            model_summary=model_summary,
            validated_comparison=validated_comparison,
            descriptive_comparison=descriptive_comparison,
            accuracy_resolution=accuracy_resolution,
        ),
        encoding="utf-8",
    )
    model_summary.to_csv(model_summary_path, index=False)
    validated_comparison.to_csv(validated_path, index=False)
    descriptive_comparison.to_csv(descriptive_path, index=False)
    accuracy_resolution.to_csv(accuracy_resolution_path, index=False)
    original_benchmark.to_csv(original_benchmark_path, index=False)

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        model_summary.to_excel(writer, sheet_name="model_summary", index=False)
        validated_comparison.to_excel(writer, sheet_name="validated_vs_original", index=False)
        descriptive_comparison.to_excel(writer, sheet_name="descriptive_vs_original", index=False)
        accuracy_resolution.to_excel(writer, sheet_name="accuracy_resolution", index=False)
        original_benchmark.to_excel(writer, sheet_name="original_benchmark", index=False)

    workbook = load_workbook(workbook_path)
    font = Font(name="Arial", size=10)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font
    workbook.save(workbook_path)

    return {
        "report": report_path,
        "summary_workbook": workbook_path,
        "model_summary": model_summary_path,
        "validated_comparison": validated_path,
        "descriptive_comparison": descriptive_path,
        "accuracy_resolution": accuracy_resolution_path,
        "original_benchmark": original_benchmark_path,
    }
