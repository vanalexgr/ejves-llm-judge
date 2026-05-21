"""Restricted Phase 6 comparator scoring pipeline."""

from __future__ import annotations

import asyncio
from dataclasses import dataclass
import json
import os
from pathlib import Path
import re
import time
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font
import pandas as pd
from dotenv import load_dotenv
from pydantic import BaseModel, ConfigDict, Field
from sklearn.metrics import cohen_kappa_score

from .judges.base import JudgeResult
from .judges.claude import ClaudeJudge
from .judges.openai import OpenAIJudge
from .prompt_builder import blind_model_identifiers, normalize_question_text
from .rubric import RUBRIC_COMPLEMENTARITY, RUBRIC_DISCERN, RUBRIC_TONE, RUBRIC_URGENCY
from .scoring import (
    ANTHROPIC_PROVIDER_CONCURRENCY,
    _aggregated_dataframe,
    _json_path,
    _raw_records_to_dataframe,
    _run_provider_pool,
    estimate_cost_usd,
)


VALID_SOURCE_MODELS = ("claude_free", "chatgpt_free", "gemini_free")
DOMAIN_TO_QUESTION_GROUP = {
    "signs_symptoms": "symptoms",
    "natural_history": "symptoms",
    "medical_advice": "treatment",
    "best_treatment": "treatment",
}
REQUIRED_INPUT_COLUMNS = (
    "response_id",
    "disease",
    "domain",
    "question_text",
    "response_text",
    "source_model",
)
VALIDATED_JUDGE_ITEMS = ("tone", "complementarity", "gilbert_urgency", "discern_q7")
DESCRIPTIVE_JUDGE_ITEMS = ("comprehensiveness", "clarity")
REVIEW_META_COLUMNS = (
    "response_id",
    "source_model",
    "disease",
    "domain",
    "question_group",
    "question_text",
    "response_text",
)

COMPARATOR_SYSTEM_PROMPT = """SYSTEM: You are a board-certified vascular surgeon evaluating an AI chatbot's
response to a patient-style question for the comparator arm of a vascular
surgery study. Accuracy is assessed separately by a human vascular surgeon and
is not part of this scoring task. Score only the domains requested in the
rubric below. Do not infer beyond what the response actually states. Output only
valid JSON matching the schema. No commentary outside the JSON."""

COMPARATOR_RUBRIC_QUALITY = """### E. Descriptive quality domains — 1 to 5 each

- **Comprehensiveness** — coverage of clinically important aspects of the question. 1 = very poor (major omissions); 5 = excellent (complete coverage)
- **Clarity** — readability and structural clarity of the response for a layperson. 1 = very poor (confusing or jargon-heavy); 5 = excellent (clear and accessible)

These two domains are recorded descriptively only in Phase 6 and are not validated primary comparator endpoints.
"""


@dataclass(frozen=True)
class ComparatorPromptBuildResult:
    prompt: str
    response_id: str
    question_group: str
    blinding_substitution_count: int
    prompt_character_length: int


@dataclass(frozen=True)
class ComparatorTarget:
    response_id: str
    source_model: str
    disease: str
    domain: str
    question_group: str
    question_text: str
    response_text: str
    prompt: str
    blinding_substitution_count: int


@dataclass(frozen=True)
class Phase6Summary:
    response_count: int
    successful_calls: int
    failed_calls: int
    failures: tuple[str, ...]
    total_latency_seconds: float
    estimated_cost_usd: float
    openai_model_requested: str
    claude_model_requested: str
    mario_review_sheet_path: Path


class ComparatorSymptomJudgeOutput(BaseModel):
    """Structured output for comparator symptom and natural-history items."""

    model_config = ConfigDict(extra="forbid")

    tone: int = Field(ge=0, le=2)
    complementarity: int = Field(ge=0, le=1)
    gilbert_urgency: int = Field(ge=0, le=4)
    comprehensiveness: int = Field(ge=1, le=5)
    clarity: int = Field(ge=1, le=5)
    rationale: str = Field(
        min_length=20,
        max_length=1500,
        description="Concise justification for the scores.",
    )


class ComparatorTreatmentJudgeOutput(BaseModel):
    """Structured output for comparator treatment items."""

    model_config = ConfigDict(extra="forbid")

    tone: int = Field(ge=0, le=2)
    complementarity: int = Field(ge=0, le=1)
    gilbert_urgency: int = Field(ge=0, le=4)
    discern_q7: int = Field(ge=1, le=5)
    comprehensiveness: int = Field(ge=1, le=5)
    clarity: int = Field(ge=1, le=5)
    rationale: str = Field(
        min_length=20,
        max_length=1500,
        description="Concise justification for the scores.",
    )


def comparator_output_model_for_question_group(question_group: str) -> type[BaseModel]:
    if question_group == "symptoms":
        return ComparatorSymptomJudgeOutput
    if question_group == "treatment":
        return ComparatorTreatmentJudgeOutput
    raise ValueError(f"Unsupported question_group: {question_group!r}")


def comparator_rubric_text_for_question_group(question_group: str) -> str:
    blocks = [RUBRIC_TONE, RUBRIC_COMPLEMENTARITY, RUBRIC_URGENCY]
    if question_group == "treatment":
        blocks.append(RUBRIC_DISCERN)
    elif question_group != "symptoms":
        raise ValueError(f"Unsupported question_group: {question_group!r}")
    blocks.append(COMPARATOR_RUBRIC_QUALITY)
    return "\n\n".join(blocks)


def comparator_output_format_template(question_group: str) -> str:
    payload: dict[str, object] = {
        "tone": 0,
        "complementarity": 0,
        "gilbert_urgency": 0,
    }
    if question_group == "treatment":
        payload["discern_q7"] = 1
    payload.update(
        {
            "comprehensiveness": 1,
            "clarity": 1,
            "rationale": "Concise justification goes here.",
        }
    )
    return json.dumps(payload, indent=2)


def build_comparator_prompt(
    *,
    response_id: str,
    question_text: str,
    response_text: str,
    question_group: str,
    domain: str,
) -> ComparatorPromptBuildResult:
    blinded = blind_model_identifiers(response_text)
    rubric_text = comparator_rubric_text_for_question_group(question_group)
    output_format = comparator_output_format_template(question_group)
    normalized_question_text = normalize_question_text(question_text)
    prompt = (
        f"{COMPARATOR_SYSTEM_PROMPT}\n\n"
        "USER:\n"
        f"QUESTION: {normalized_question_text}\n\n"
        "RESPONSE TO SCORE:\n"
        "---\n"
        f"{blinded.text}\n"
        "---\n\n"
        f"QUESTION CATEGORY: {question_group}  ({domain})\n\n"
        "RUBRIC:\n"
        f"{rubric_text}\n\n"
        "Accuracy is intentionally omitted from this task because it is assigned separately by a "
        "human vascular surgeon.\n"
        "Replace each numeric placeholder with your integer score (within the stated range). "
        "Replace rationale with a concise justification (roughly 2–4 sentences).\n\n"
        "OUTPUT FORMAT (JSON):\n"
        f"{output_format}\n"
    )
    return ComparatorPromptBuildResult(
        prompt=prompt,
        response_id=response_id,
        question_group=question_group,
        blinding_substitution_count=blinded.substitution_count,
        prompt_character_length=len(prompt),
    )


def load_comparator_input(input_csv: Path) -> pd.DataFrame:
    frame = pd.read_csv(input_csv)
    missing_columns = [column for column in REQUIRED_INPUT_COLUMNS if column not in frame.columns]
    if missing_columns:
        raise ValueError(f"Missing required comparator columns: {missing_columns}")

    data = frame.copy()
    for column in REQUIRED_INPUT_COLUMNS:
        data[column] = data[column].astype(str).str.strip()
        if data[column].eq("").any():
            raise ValueError(f"Comparator column {column!r} contains blank values.")

    if data["response_id"].duplicated().any():
        duplicates = data.loc[data["response_id"].duplicated(keep=False), "response_id"].tolist()
        raise ValueError(f"Comparator response_id values must be unique. Duplicates: {duplicates}")

    data["source_model"] = data["source_model"].str.lower()
    invalid_source_models = sorted(set(data["source_model"]) - set(VALID_SOURCE_MODELS))
    if invalid_source_models:
        raise ValueError(
            f"Unsupported source_model values: {invalid_source_models}. "
            f"Expected one of {VALID_SOURCE_MODELS}."
        )

    data["domain"] = data["domain"].str.lower()
    invalid_domains = sorted(set(data["domain"]) - set(DOMAIN_TO_QUESTION_GROUP))
    if invalid_domains:
        raise ValueError(
            f"Unsupported domain values: {invalid_domains}. "
            f"Expected one of {sorted(DOMAIN_TO_QUESTION_GROUP)}."
        )

    data["question_group"] = data["domain"].map(DOMAIN_TO_QUESTION_GROUP)
    data = data.sort_values(["source_model", "response_id"], kind="stable").reset_index(drop=True)
    return data


def build_comparator_targets(input_frame: pd.DataFrame) -> list[ComparatorTarget]:
    targets: list[ComparatorTarget] = []
    for row in input_frame.itertuples(index=False):
        prompt_result = build_comparator_prompt(
            response_id=row.response_id,
            question_text=row.question_text,
            response_text=row.response_text,
            question_group=row.question_group,
            domain=row.domain,
        )
        targets.append(
            ComparatorTarget(
                response_id=row.response_id,
                source_model=row.source_model,
                disease=row.disease,
                domain=row.domain,
                question_group=row.question_group,
                question_text=row.question_text,
                response_text=row.response_text,
                prompt=prompt_result.prompt,
                blinding_substitution_count=prompt_result.blinding_substitution_count,
            )
        )
    return targets


def comparator_target_frame(targets: list[ComparatorTarget]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "response_id": target.response_id,
                "source_model": target.source_model,
                "disease": target.disease,
                "domain": target.domain,
                "question_group": target.question_group,
                "question_text": target.question_text,
                "response_text": target.response_text,
                "blinding_substitution_count": target.blinding_substitution_count,
            }
            for target in targets
        ]
    ).sort_values(["source_model", "response_id"], kind="stable").reset_index(drop=True)


class _ComparatorJudgeMixin:
    def score(self, *, response_id: str, prompt: str, question_group: str) -> JudgeResult:
        model_cls = comparator_output_model_for_question_group(question_group)
        started = time.perf_counter()
        raw_response, request_id, model_used, request_payload, parsed = self._call_with_retries(
            response_id=response_id,
            prompt=prompt,
            model_cls=model_cls,
        )
        latency = time.perf_counter() - started
        validated = model_cls.model_validate(parsed)
        return JudgeResult(
            judge_name=self.judge_name,
            response_id=response_id,
            question_group=question_group,
            model_used=model_used,
            latency_seconds=latency,
            request_id=request_id,
            request_payload=request_payload,
            raw_response=raw_response,
            parsed_response=validated,
        )

    def build_logged_request_payload(self, *, prompt: str, question_group: str) -> dict[str, Any]:
        model_cls = comparator_output_model_for_question_group(question_group)
        return self._build_logged_request_payload(prompt=prompt, model_cls=model_cls)


class ComparatorClaudeJudge(_ComparatorJudgeMixin, ClaudeJudge):
    """Claude wrapper using the restricted Phase 6 schema."""


class ComparatorOpenAIJudge(_ComparatorJudgeMixin, OpenAIJudge):
    """OpenAI wrapper using the restricted Phase 6 schema."""


def build_mario_accuracy_review_sheet(input_frame: pd.DataFrame) -> pd.DataFrame:
    review = input_frame[list(REVIEW_META_COLUMNS)].copy()
    review["mario_accuracy"] = pd.Series([pd.NA] * len(review), dtype="Int64")
    review["mario_accuracy_notes"] = pd.NA
    review["accuracy_reviewer"] = "Mario"
    review["accuracy_role"] = "validated_endpoint"
    return review


def _read_review_table(review_path: Path) -> pd.DataFrame:
    suffix = review_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        raw = pd.read_excel(review_path, sheet_name=0, header=None, dtype=object)
    elif suffix in {".csv", ".tsv"}:
        separator = "\t" if suffix == ".tsv" else ","
        raw = pd.read_csv(review_path, header=None, dtype=object, sep=separator)
    else:
        raise ValueError(f"Unsupported review file type: {review_path.suffix}")

    raw = raw.dropna(how="all").reset_index(drop=True)
    if raw.empty:
        raise ValueError(f"Review file is empty: {review_path}")

    header_row_index = None
    normalized = raw.fillna("").map(lambda value: str(value).strip())
    for index, row in normalized.iterrows():
        row_values = {value.lower() for value in row.tolist() if value}
        if "response_id" in row_values:
            header_row_index = index
            break
    if header_row_index is None:
        raise ValueError(f"Could not locate a header row containing response_id in {review_path}")

    header = normalized.iloc[header_row_index].tolist()
    data = raw.iloc[header_row_index + 1 :].reset_index(drop=True)
    data.columns = header
    keep_columns = [
        column
        for column in data.columns
        if str(column).strip() and not str(column).lower().startswith("unnamed")
    ]
    data = data[keep_columns].copy()
    data.columns = [str(column).strip() for column in data.columns]
    data = data.dropna(how="all").reset_index(drop=True)
    return data


def load_accuracy_review(
    review_path: Path,
    *,
    score_column: str,
    notes_column: str,
    expected_response_ids: set[str],
) -> pd.DataFrame:
    review = _read_review_table(review_path)
    required_columns = {"response_id", score_column}
    missing = required_columns - set(review.columns)
    if missing:
        raise ValueError(f"Missing required review columns in {review_path.name}: {sorted(missing)}")

    review = review.copy()
    review["response_id"] = review["response_id"].astype(str).str.strip()
    missing_ids = sorted(expected_response_ids - set(review["response_id"]))
    extra_ids = sorted(set(review["response_id"]) - expected_response_ids)
    if missing_ids or extra_ids:
        raise ValueError(
            f"Review response_id mismatch for {review_path.name}. Missing={missing_ids}, extra={extra_ids}"
        )

    review[score_column] = pd.to_numeric(review[score_column], errors="coerce").astype("Int64")
    invalid_scores = review.loc[
        review[score_column].notna() & ~review[score_column].between(1, 5),
        score_column,
    ]
    if not invalid_scores.empty:
        raise ValueError(
            f"Accuracy scores in {review_path.name} must be integers 1-5. Invalid values: {invalid_scores.tolist()}"
        )
    if notes_column not in review.columns:
        review[notes_column] = pd.NA
    review[notes_column] = review[notes_column].where(review[notes_column].notna(), pd.NA)
    return review[["response_id", score_column, notes_column]].copy()


def load_mario_accuracy_review(review_csv: Path, *, expected_response_ids: set[str]) -> pd.DataFrame:
    return load_accuracy_review(
        review_csv,
        score_column="mario_accuracy",
        notes_column="mario_accuracy_notes",
        expected_response_ids=expected_response_ids,
    )


def load_vga_accuracy_review(review_path: Path, *, expected_response_ids: set[str]) -> pd.DataFrame:
    return load_accuracy_review(
        review_path,
        score_column="vga_accuracy",
        notes_column="vga_accuracy_notes",
        expected_response_ids=expected_response_ids,
    )


def load_third_accuracy_review(review_path: Path, *, expected_response_ids: set[str]) -> pd.DataFrame:
    return load_accuracy_review(
        review_path,
        score_column="accuracy_score",
        notes_column="rationale",
        expected_response_ids=expected_response_ids,
    ).rename(
        columns={
            "accuracy_score": "third_reviewer_accuracy",
            "rationale": "third_reviewer_notes",
        }
    )


def load_accuracy_rationale(
    review_path: Path,
    *,
    score_column: str,
    rationale_column: str,
    expected_response_ids: set[str],
) -> pd.DataFrame:
    review = _read_review_table(review_path)
    required_columns = {"response_id", score_column, rationale_column}
    missing = required_columns - set(review.columns)
    if missing:
        raise ValueError(f"Missing required rationale columns in {review_path.name}: {sorted(missing)}")

    review = review.copy()
    review["response_id"] = review["response_id"].astype(str).str.strip()
    missing_ids = sorted(expected_response_ids - set(review["response_id"]))
    extra_ids = sorted(set(review["response_id"]) - expected_response_ids)
    if missing_ids or extra_ids:
        raise ValueError(
            f"Rationale response_id mismatch for {review_path.name}. Missing={missing_ids}, extra={extra_ids}"
        )

    review[score_column] = pd.to_numeric(review[score_column], errors="coerce").astype("Int64")
    invalid_scores = review.loc[
        review[score_column].notna() & ~review[score_column].between(1, 5),
        score_column,
    ]
    if not invalid_scores.empty:
        raise ValueError(
            f"Rationale scores in {review_path.name} must be integers 1-5. Invalid values: {invalid_scores.tolist()}"
        )

    rationale = review[rationale_column].astype(str).str.strip()
    review[rationale_column] = rationale.where(rationale.ne(""), pd.NA)
    return review[["response_id", score_column, rationale_column]].copy()


def load_mario_accuracy_rationale(
    review_path: Path,
    *,
    expected_response_ids: set[str],
) -> pd.DataFrame:
    return load_accuracy_rationale(
        review_path,
        score_column="your_score",
        rationale_column="your_rationale",
        expected_response_ids=expected_response_ids,
    ).rename(
        columns={
            "your_score": "mario_rationale_score",
            "your_rationale": "mario_rationale",
        }
    )


def load_vga_accuracy_rationale(
    review_path: Path,
    *,
    expected_response_ids: set[str],
) -> pd.DataFrame:
    return load_accuracy_rationale(
        review_path,
        score_column="your_score",
        rationale_column="your_rationale",
        expected_response_ids=expected_response_ids,
    ).rename(
        columns={
            "your_score": "vga_rationale_score",
            "your_rationale": "vga_rationale",
        }
    )


def extract_requested_revised_accuracy(rationale: object) -> int | None:
    if rationale is None or pd.isna(rationale):
        return None
    text = str(rationale).strip().lower()
    if not text:
        return None
    patterns = (
        r"\brate again as\s*([1-5])\b",
        r"\brerate(?: it)? as\s*([1-5])\b",
        r"\bre[- ]?rate(?: it)? as\s*([1-5])\b",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return int(match.group(1))
    return None


def build_dual_accuracy_review_table(
    *,
    reference_frame: pd.DataFrame,
    mario_review: pd.DataFrame,
    vga_review: pd.DataFrame,
) -> pd.DataFrame:
    merged = (
        reference_frame[list(REVIEW_META_COLUMNS)]
        .drop_duplicates(subset=["response_id"])
        .merge(mario_review, on="response_id", how="left", validate="one_to_one")
        .merge(vga_review, on="response_id", how="left", validate="one_to_one")
        .sort_values(["source_model", "response_id"], kind="stable")
        .reset_index(drop=True)
    )
    merged["human_accuracy_mean"] = (
        merged[["mario_accuracy", "vga_accuracy"]].astype(float).mean(axis=1)
    )
    merged["human_accuracy_abs_diff"] = (
        merged["mario_accuracy"].astype(float) - merged["vga_accuracy"].astype(float)
    ).abs()
    merged["human_accuracy_resolution_status"] = "minor_disagreement"
    merged.loc[merged["human_accuracy_abs_diff"].eq(0), "human_accuracy_resolution_status"] = (
        "exact_agreement"
    )
    merged.loc[merged["human_accuracy_abs_diff"].ge(2), "human_accuracy_resolution_status"] = (
        "adjudication_needed"
    )
    merged["validated_accuracy"] = merged["human_accuracy_mean"].where(
        merged["human_accuracy_abs_diff"].le(1),
        pd.NA,
    )
    return merged


def build_accuracy_adjudication_table(
    *,
    dual_accuracy_reviews: pd.DataFrame,
    mario_rationale: pd.DataFrame,
    vga_rationale: pd.DataFrame,
) -> pd.DataFrame:
    adjudication = dual_accuracy_reviews.loc[
        dual_accuracy_reviews["human_accuracy_resolution_status"].eq("adjudication_needed")
    ].copy()
    adjudication = adjudication.merge(
        mario_rationale,
        on="response_id",
        how="left",
        validate="one_to_one",
    ).merge(
        vga_rationale,
        on="response_id",
        how="left",
        validate="one_to_one",
    )
    adjudication["mario_requested_revised_accuracy"] = adjudication["mario_rationale"].map(
        extract_requested_revised_accuracy
    )
    adjudication["mario_requested_revised_accuracy"] = adjudication[
        "mario_requested_revised_accuracy"
    ].astype("Int64")
    adjudication["mario_effective_accuracy"] = adjudication["mario_requested_revised_accuracy"].combine_first(
        adjudication["mario_accuracy"]
    )
    adjudication["mario_revision_applied"] = adjudication["mario_requested_revised_accuracy"].notna()
    adjudication["effective_abs_diff_after_mario_revision"] = (
        adjudication["mario_effective_accuracy"].astype(float) - adjudication["vga_accuracy"].astype(float)
    ).abs()
    adjudication["would_resolve_after_mario_revision"] = adjudication[
        "effective_abs_diff_after_mario_revision"
    ].le(1)
    adjudication = adjudication.sort_values(["source_model", "response_id"], kind="stable").reset_index(drop=True)
    adjudication.insert(
        0,
        "adjudication_case_id",
        [f"ADJ_{index:02d}" for index in range(1, len(adjudication) + 1)],
    )
    adjudication["third_reviewer_accuracy"] = pd.Series([pd.NA] * len(adjudication), dtype="Int64")
    adjudication["third_reviewer_notes"] = pd.NA
    return adjudication


def build_revised_dual_accuracy_review_table(
    dual_accuracy_reviews: pd.DataFrame,
    adjudication_table: pd.DataFrame,
) -> pd.DataFrame:
    revised = dual_accuracy_reviews.copy()
    revised["mario_accuracy_original"] = revised["mario_accuracy"]
    revised = revised.merge(
        adjudication_table[
            ["response_id", "mario_requested_revised_accuracy", "mario_rationale", "vga_rationale"]
        ],
        on="response_id",
        how="left",
    )
    revised["mario_revision_applied"] = revised["mario_requested_revised_accuracy"].notna()
    revised["mario_accuracy"] = revised["mario_requested_revised_accuracy"].combine_first(
        revised["mario_accuracy"]
    ).astype("Int64")
    revised["human_accuracy_mean"] = revised[["mario_accuracy", "vga_accuracy"]].astype(float).mean(axis=1)
    revised["human_accuracy_abs_diff"] = (
        revised["mario_accuracy"].astype(float) - revised["vga_accuracy"].astype(float)
    ).abs()
    revised["human_accuracy_resolution_status"] = "minor_disagreement"
    revised.loc[revised["human_accuracy_abs_diff"].eq(0), "human_accuracy_resolution_status"] = (
        "exact_agreement"
    )
    revised.loc[revised["human_accuracy_abs_diff"].ge(2), "human_accuracy_resolution_status"] = (
        "adjudication_needed"
    )
    revised["validated_accuracy"] = revised["human_accuracy_mean"].where(
        revised["human_accuracy_abs_diff"].le(1),
        pd.NA,
    )
    return revised.sort_values(["source_model", "response_id"], kind="stable").reset_index(drop=True)


def build_final_accuracy_review_table(
    *,
    revised_dual_accuracy_reviews: pd.DataFrame,
    adjudication_table: pd.DataFrame,
    third_reviewer_review: pd.DataFrame,
) -> pd.DataFrame:
    final = revised_dual_accuracy_reviews.copy()
    final["mario_accuracy_final"] = final["mario_accuracy"]
    final["vga_accuracy_final"] = final["vga_accuracy"]
    final = final.merge(
        adjudication_table[["response_id", "adjudication_case_id"]],
        on="response_id",
        how="left",
    ).merge(
        third_reviewer_review,
        on="response_id",
        how="left",
    )

    adjudication_ids = set(adjudication_table["response_id"])
    final["third_reviewer_used"] = final["response_id"].isin(adjudication_ids)
    missing_third = final.loc[
        final["third_reviewer_used"] & final["third_reviewer_accuracy"].isna(),
        "response_id",
    ].tolist()
    if missing_third:
        raise ValueError(f"Missing third reviewer scores for adjudication responses: {missing_third}")

    final["final_validated_accuracy"] = final["validated_accuracy"].astype(float)
    final["final_accuracy_resolution_status"] = final["human_accuracy_resolution_status"]
    final["final_accuracy_source"] = final["human_accuracy_resolution_status"].map(
        {
            "exact_agreement": "dual_surgeon_agreement",
            "minor_disagreement": "dual_surgeon_mean",
            "adjudication_needed": "dual_surgeon_adjudication_needed",
        }
    )
    adjudication_mask = final["third_reviewer_used"]
    if adjudication_mask.any():
        final.loc[adjudication_mask, "final_validated_accuracy"] = final.loc[
            adjudication_mask,
            ["mario_accuracy_final", "vga_accuracy_final", "third_reviewer_accuracy"],
        ].astype(float).median(axis=1)
        final.loc[adjudication_mask, "final_accuracy_resolution_status"] = "triple_surgeon_resolved"
        final.loc[adjudication_mask, "final_accuracy_source"] = "triple_surgeon_median"

    return final.sort_values(["source_model", "response_id"], kind="stable").reset_index(drop=True)


def build_dual_accuracy_summary(merged_reviews: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    scores = merged_reviews[["mario_accuracy", "vga_accuracy"]].astype(int)
    summary_rows = [
        {"metric": "total_responses", "value": int(len(merged_reviews))},
        {
            "metric": "exact_agreements",
            "value": int(merged_reviews["human_accuracy_resolution_status"].eq("exact_agreement").sum()),
        },
        {
            "metric": "minor_disagreements",
            "value": int(merged_reviews["human_accuracy_resolution_status"].eq("minor_disagreement").sum()),
        },
        {
            "metric": "adjudication_needed",
            "value": int(merged_reviews["human_accuracy_resolution_status"].eq("adjudication_needed").sum()),
        },
        {"metric": "mean_mario_accuracy", "value": float(merged_reviews["mario_accuracy"].mean())},
        {"metric": "mean_vga_accuracy", "value": float(merged_reviews["vga_accuracy"].mean())},
        {
            "metric": "mean_abs_difference",
            "value": float(merged_reviews["human_accuracy_abs_diff"].mean()),
        },
        {
            "metric": "quadratic_weighted_kappa",
            "value": float(cohen_kappa_score(scores["mario_accuracy"], scores["vga_accuracy"], weights="quadratic")),
        },
        {
            "metric": "unweighted_kappa",
            "value": float(cohen_kappa_score(scores["mario_accuracy"], scores["vga_accuracy"])),
        },
    ]
    summary = pd.DataFrame(summary_rows)

    by_model = (
        merged_reviews.groupby("source_model", sort=True)
        .agg(
            response_count=("response_id", "size"),
            mean_mario_accuracy=("mario_accuracy", "mean"),
            mean_vga_accuracy=("vga_accuracy", "mean"),
            mean_abs_difference=("human_accuracy_abs_diff", "mean"),
            exact_agreements=("human_accuracy_resolution_status", lambda values: int((values == "exact_agreement").sum())),
            minor_disagreements=("human_accuracy_resolution_status", lambda values: int((values == "minor_disagreement").sum())),
            adjudication_needed=("human_accuracy_resolution_status", lambda values: int((values == "adjudication_needed").sum())),
        )
        .reset_index()
    )
    return summary, by_model


def build_final_accuracy_summary(final_reviews: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary_rows = [
        {"metric": "total_responses", "value": int(len(final_reviews))},
        {
            "metric": "triple_surgeon_cases",
            "value": int(final_reviews["third_reviewer_used"].fillna(False).sum()),
        },
        {
            "metric": "dual_exact_agreements",
            "value": int(final_reviews["human_accuracy_resolution_status"].eq("exact_agreement").sum()),
        },
        {
            "metric": "dual_minor_disagreements",
            "value": int(final_reviews["human_accuracy_resolution_status"].eq("minor_disagreement").sum()),
        },
        {
            "metric": "mean_final_validated_accuracy",
            "value": float(final_reviews["final_validated_accuracy"].astype(float).mean()),
        },
    ]
    summary = pd.DataFrame(summary_rows)
    by_model = (
        final_reviews.groupby("source_model", sort=True)
        .agg(
            response_count=("response_id", "size"),
            triple_surgeon_cases=("third_reviewer_used", lambda values: int(pd.Series(values).fillna(False).sum())),
            mean_final_validated_accuracy=("final_validated_accuracy", "mean"),
        )
        .reset_index()
    )
    return summary, by_model


def build_phase6_results_with_dual_accuracy(
    *,
    ensemble_scores: pd.DataFrame,
    dual_accuracy_reviews: pd.DataFrame,
) -> pd.DataFrame:
    results = ensemble_scores.copy()
    dual_keep = [
        "response_id",
        "mario_accuracy",
        "mario_accuracy_notes",
        "vga_accuracy",
        "vga_accuracy_notes",
        "human_accuracy_mean",
        "human_accuracy_abs_diff",
        "human_accuracy_resolution_status",
        "validated_accuracy",
    ]
    results = results.merge(
        dual_accuracy_reviews[dual_keep],
        on="response_id",
        how="left",
        validate="one_to_one",
    )
    results["accuracy_source"] = results["human_accuracy_resolution_status"].map(
        {
            "exact_agreement": "dual_surgeon_agreement",
            "minor_disagreement": "dual_surgeon_mean",
            "adjudication_needed": "dual_surgeon_adjudication_needed",
        }
    )
    results["validated_tone"] = results["tone"]
    results["validated_complementarity"] = results["complementarity"]
    results["validated_gilbert_urgency"] = results["gilbert_urgency"]
    results["validated_discern_q7"] = results["discern_q7"]
    results["descriptive_comprehensiveness"] = results["comprehensiveness"]
    results["descriptive_clarity"] = results["clarity"]
    return results


def build_phase6_results_with_final_accuracy(
    *,
    ensemble_scores: pd.DataFrame,
    final_accuracy_reviews: pd.DataFrame,
) -> pd.DataFrame:
    results = ensemble_scores.copy()
    keep_columns = [
        "response_id",
        "mario_accuracy_original",
        "mario_requested_revised_accuracy",
        "mario_accuracy_final",
        "mario_accuracy_notes",
        "mario_rationale",
        "vga_accuracy_final",
        "vga_accuracy_notes",
        "vga_rationale",
        "third_reviewer_accuracy",
        "third_reviewer_notes",
        "third_reviewer_used",
        "adjudication_case_id",
        "human_accuracy_mean",
        "human_accuracy_abs_diff",
        "human_accuracy_resolution_status",
        "final_validated_accuracy",
        "final_accuracy_resolution_status",
        "final_accuracy_source",
    ]
    results = results.merge(
        final_accuracy_reviews[keep_columns],
        on="response_id",
        how="left",
        validate="one_to_one",
    )
    results["validated_accuracy"] = results["final_validated_accuracy"]
    results["accuracy_source"] = results["final_accuracy_source"]
    results["validated_tone"] = results["tone"]
    results["validated_complementarity"] = results["complementarity"]
    results["validated_gilbert_urgency"] = results["gilbert_urgency"]
    results["validated_discern_q7"] = results["discern_q7"]
    results["descriptive_comprehensiveness"] = results["comprehensiveness"]
    results["descriptive_clarity"] = results["clarity"]
    return results


def write_dual_accuracy_workbook(
    *,
    summary: pd.DataFrame,
    by_model: pd.DataFrame,
    merged_reviews: pd.DataFrame,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    adjudication = merged_reviews.loc[
        merged_reviews["human_accuracy_resolution_status"].eq("adjudication_needed")
    ].copy()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        by_model.to_excel(writer, sheet_name="by_model", index=False)
        merged_reviews.to_excel(writer, sheet_name="merged_reviews", index=False)
        adjudication.to_excel(writer, sheet_name="adjudication_needed", index=False)

    workbook = load_workbook(output_path)
    font = Font(name="Arial", size=10)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font
    workbook.save(output_path)


def build_blinded_third_reviewer_packet(adjudication_table: pd.DataFrame) -> pd.DataFrame:
    packet = adjudication_table[
        [
            "adjudication_case_id",
            "disease",
            "domain",
            "question_group",
            "question_text",
            "response_text",
        ]
    ].copy()
    packet["accuracy_scale"] = (
        "1 = clinically misleading or major errors | "
        "2 = defensible but significant gaps | "
        "3 = defensible with minor gaps | "
        "4 = accurate with minor omissions | "
        "5 = accurate, balanced, patient-appropriate"
    )
    packet["third_reviewer_accuracy"] = pd.Series([pd.NA] * len(packet), dtype="Int64")
    packet["third_reviewer_notes"] = pd.NA
    return packet


def write_accuracy_adjudication_workbook(
    *,
    adjudication_table: pd.DataFrame,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    blinded_packet = build_blinded_third_reviewer_packet(adjudication_table)
    instructions = pd.DataFrame(
        [
            {
                "instruction": (
                    "Score only clinical accuracy on a 1-5 scale. Cases are blinded to source model."
                )
            },
            {
                "instruction": (
                    "Use third_reviewer_accuracy for the integer score and third_reviewer_notes for brief rationale if helpful."
                )
            },
        ]
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="instructions", index=False)
        adjudication_table.to_excel(writer, sheet_name="master_with_mapping", index=False)
        blinded_packet.to_excel(writer, sheet_name="third_reviewer_blinded", index=False)

    workbook = load_workbook(output_path)
    font = Font(name="Arial", size=10)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font
    workbook.save(output_path)


def write_third_reviewer_packet_workbook(
    *,
    third_reviewer_packet: pd.DataFrame,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    instructions = pd.DataFrame(
        [
            {
                "instruction": (
                    "Score only clinical accuracy on a 1-5 scale. Cases are blinded to source model."
                )
            },
            {
                "instruction": (
                    "Use third_reviewer_accuracy for the integer score and third_reviewer_notes for brief rationale if helpful."
                )
            },
        ]
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="instructions", index=False)
        third_reviewer_packet.to_excel(writer, sheet_name="third_reviewer_packet", index=False)

    workbook = load_workbook(output_path)
    font = Font(name="Arial", size=10)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font
    workbook.save(output_path)


def build_phase6_ensemble_scores(
    judge_aggregated: pd.DataFrame,
    target_meta: pd.DataFrame,
) -> pd.DataFrame:
    meta = target_meta.set_index("response_id")
    score_items = ("tone", "complementarity", "gilbert_urgency", "discern_q7", "comprehensiveness", "clarity")
    rows: list[dict[str, Any]] = []
    for response_id, group in judge_aggregated.groupby("response_id", sort=True):
        row: dict[str, Any] = {
            "response_id": response_id,
            "source_model": meta.loc[response_id, "source_model"],
            "disease": meta.loc[response_id, "disease"],
            "domain": meta.loc[response_id, "domain"],
            "question_group": meta.loc[response_id, "question_group"],
            "question_text": meta.loc[response_id, "question_text"],
            "response_text": meta.loc[response_id, "response_text"],
        }
        for item in score_items:
            values = group[f"{item}_aggregated"].dropna().astype(float)
            row[item] = float(values.mean()) if not values.empty else pd.NA
        rows.append(row)
    return pd.DataFrame(rows).sort_values(["source_model", "response_id"], kind="stable").reset_index(drop=True)


def build_phase6_results(
    *,
    ensemble_scores: pd.DataFrame,
    mario_review: pd.DataFrame | None,
) -> pd.DataFrame:
    results = ensemble_scores.copy()
    results["accuracy_source"] = "mario_review_pending"
    results["mario_accuracy"] = pd.Series([pd.NA] * len(results), dtype="Int64")
    results["mario_accuracy_notes"] = pd.NA
    if mario_review is not None:
        keep_columns = ["response_id", "mario_accuracy"]
        if "mario_accuracy_notes" in mario_review.columns:
            keep_columns.append("mario_accuracy_notes")
        results = results.drop(columns=["mario_accuracy", "mario_accuracy_notes"]).merge(
            mario_review[keep_columns],
            on="response_id",
            how="left",
        )
        results["accuracy_source"] = "Mario"

    results["validated_tone"] = results["tone"]
    results["validated_complementarity"] = results["complementarity"]
    results["validated_gilbert_urgency"] = results["gilbert_urgency"]
    results["validated_discern_q7"] = results["discern_q7"]
    results["descriptive_comprehensiveness"] = results["comprehensiveness"]
    results["descriptive_clarity"] = results["clarity"]
    return results


def write_source_model_slices(
    *,
    raw_df: pd.DataFrame,
    judge_aggregated: pd.DataFrame,
    ensemble_scores: pd.DataFrame,
    phase6_results: pd.DataFrame,
    output_dir: Path,
) -> None:
    by_model_dir = output_dir / "by_source_model"
    by_model_dir.mkdir(parents=True, exist_ok=True)
    for source_model in sorted(ensemble_scores["source_model"].dropna().unique()):
        model_dir = by_model_dir / source_model
        model_dir.mkdir(parents=True, exist_ok=True)
        response_ids = set(ensemble_scores.loc[ensemble_scores["source_model"].eq(source_model), "response_id"])
        raw_df.loc[raw_df["response_id"].isin(response_ids)].to_parquet(
            model_dir / "judge_scores_raw.parquet",
            index=False,
        )
        judge_aggregated.loc[judge_aggregated["response_id"].isin(response_ids)].to_parquet(
            model_dir / "judge_scores_aggregated.parquet",
            index=False,
        )
        ensemble_scores.loc[ensemble_scores["source_model"].eq(source_model)].to_parquet(
            model_dir / "ensemble_scores.parquet",
            index=False,
        )
        phase6_results.loc[phase6_results["source_model"].eq(source_model)].to_parquet(
            model_dir / "comparator_results.parquet",
            index=False,
        )


async def run_phase6_async(
    *,
    input_csv: Path,
    per_call_output_dir: Path,
    output_dir: Path,
    provider_concurrency: int = 3,
    mario_review_output_path: Path | None = None,
    mario_review_input_path: Path | None = None,
) -> Phase6Summary:
    load_dotenv(Path.cwd() / ".env")
    anthropic_api_key = os.getenv("ANTHROPIC_API_KEY")
    openai_api_key = os.getenv("OPENAI_API_KEY")
    if not anthropic_api_key:
        raise ValueError("Missing ANTHROPIC_API_KEY.")
    if not openai_api_key:
        raise ValueError("Missing OPENAI_API_KEY.")

    input_frame = load_comparator_input(input_csv)
    targets = build_comparator_targets(input_frame)
    target_meta = comparator_target_frame(targets)

    per_call_output_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    openai_model = ComparatorOpenAIJudge.discover_gpt5_model(api_key=openai_api_key)
    claude_judge = ComparatorClaudeJudge(api_key=anthropic_api_key, model="claude-opus-4-7")
    openai_judge = ComparatorOpenAIJudge(api_key=openai_api_key, model=openai_model)
    claude_concurrency = max(1, min(provider_concurrency, ANTHROPIC_PROVIDER_CONCURRENCY))

    provider_results = await asyncio.gather(
        _run_provider_pool(
            judge=claude_judge,
            targets=targets,
            output_dir=per_call_output_dir,
            concurrency=claude_concurrency,
        ),
        _run_provider_pool(
            judge=openai_judge,
            targets=targets,
            output_dir=per_call_output_dir,
            concurrency=provider_concurrency,
        ),
    )
    records = [payload for provider_batch in provider_results for payload in provider_batch]
    expected_calls = len(targets) * 2 * 3
    if len(records) != expected_calls:
        raise ValueError(f"Expected {expected_calls} call records, found {len(records)}.")

    for record in records:
        record["artifact_path"] = str(
            _json_path(
                per_call_output_dir,
                record["judge_name"],
                record["response_id"],
                int(record["run_index"]),
            )
        )

    raw_df = _raw_records_to_dataframe(records).merge(
        target_meta,
        on=["response_id", "question_group"],
        how="left",
    )
    judge_aggregated = _aggregated_dataframe(raw_df).merge(
        target_meta,
        on=["response_id", "question_group"],
        how="left",
    )
    ensemble_scores = build_phase6_ensemble_scores(judge_aggregated, target_meta)

    mario_review_sheet = build_mario_accuracy_review_sheet(input_frame)
    mario_review_path = mario_review_output_path or (output_dir / "mario_accuracy_review.csv")
    mario_review_sheet.to_csv(mario_review_path, index=False)

    mario_review_complete = None
    if mario_review_input_path is not None:
        mario_review_complete = load_mario_accuracy_review(
            mario_review_input_path,
            expected_response_ids=set(input_frame["response_id"]),
        )
    phase6_results = build_phase6_results(
        ensemble_scores=ensemble_scores,
        mario_review=mario_review_complete,
    )

    raw_path = output_dir / "judge_scores_raw.parquet"
    aggregated_path = output_dir / "judge_scores_aggregated.parquet"
    ensemble_path = output_dir / "ensemble_scores.parquet"
    pending_results_path = output_dir / "comparator_results.parquet"
    raw_df.to_parquet(raw_path, index=False)
    judge_aggregated.to_parquet(aggregated_path, index=False)
    ensemble_scores.to_parquet(ensemble_path, index=False)
    phase6_results.to_parquet(pending_results_path, index=False)
    write_source_model_slices(
        raw_df=raw_df,
        judge_aggregated=judge_aggregated,
        ensemble_scores=ensemble_scores,
        phase6_results=phase6_results,
        output_dir=output_dir,
    )

    successful_calls = int(raw_df["status"].eq("success").sum())
    failed_rows = raw_df.loc[raw_df["status"] != "success", ["judge_name", "response_id", "run_index"]]
    failures = tuple(
        f"{row.judge_name}:{row.response_id}:run{row.run_index}"
        for row in failed_rows.itertuples(index=False)
    )
    total_latency = float(raw_df["latency_seconds"].fillna(0).sum())
    estimated_cost = estimate_cost_usd(records)

    return Phase6Summary(
        response_count=len(targets),
        successful_calls=successful_calls,
        failed_calls=len(failures),
        failures=failures,
        total_latency_seconds=total_latency,
        estimated_cost_usd=estimated_cost,
        openai_model_requested=openai_model,
        claude_model_requested="claude-opus-4-7",
        mario_review_sheet_path=mario_review_path,
    )


def prepare_phase6_inputs(
    *,
    input_csv: Path,
    output_dir: Path,
    mario_review_output_path: Path | None = None,
) -> Path:
    input_frame = load_comparator_input(input_csv)
    output_dir.mkdir(parents=True, exist_ok=True)
    mario_review_sheet = build_mario_accuracy_review_sheet(input_frame)
    mario_review_path = mario_review_output_path or (output_dir / "mario_accuracy_review.csv")
    mario_review_sheet.to_csv(mario_review_path, index=False)
    return mario_review_path


def run_phase6(
    *,
    input_csv: Path,
    per_call_output_dir: Path,
    output_dir: Path,
    provider_concurrency: int = 3,
    mario_review_output_path: Path | None = None,
    mario_review_input_path: Path | None = None,
) -> Phase6Summary:
    return asyncio.run(
        run_phase6_async(
            input_csv=input_csv,
            per_call_output_dir=per_call_output_dir,
            output_dir=output_dir,
            provider_concurrency=provider_concurrency,
            mario_review_output_path=mario_review_output_path,
            mario_review_input_path=mario_review_input_path,
        )
    )
